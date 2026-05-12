import io
import tempfile

import pytesseract
from PIL import Image
from pdf2image import convert_from_bytes
from PyPDF2 import PdfReader


def extract_text_from_pdf(file_data: bytes) -> str:
    """Extract text from PDF. Try PyPDF2 first (fast), fall back to Tesseract OCR."""
    text = _extract_pdf_text(file_data)
    if text and len(text.strip()) > 50:
        return text.strip()
    return _ocr_pdf(file_data)


def extract_text_from_image(file_data: bytes) -> str:
    """OCR an image file."""
    image = Image.open(io.BytesIO(file_data))
    return pytesseract.image_to_string(image, lang="eng").strip()


def extract_text_from_docx(file_data: bytes) -> str:
    """Extract text from a DOCX file."""
    import docx
    doc = docx.Document(io.BytesIO(file_data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_text_from_xlsx(file_data: bytes) -> str:
    """Extract text from an XLSX file."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def extract_text(file_data: bytes, mime_type: str, filename: str = "") -> str:
    """Route to the correct extractor based on MIME type."""
    if mime_type == "application/pdf":
        return extract_text_from_pdf(file_data)
    elif mime_type in ("image/png", "image/jpeg", "image/tiff", "image/bmp", "image/webp"):
        return extract_text_from_image(file_data)
    elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return extract_text_from_docx(file_data)
    elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return extract_text_from_xlsx(file_data)
    elif mime_type.startswith("text/"):
        return file_data.decode("utf-8", errors="replace")
    return ""


def _extract_pdf_text(file_data: bytes) -> str:
    """Try extracting embedded text from PDF."""
    try:
        reader = PdfReader(io.BytesIO(file_data))
        pages = []
        for page in reader.pages[:20]:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages)
    except Exception:
        return ""


def _ocr_pdf(file_data: bytes) -> str:
    """Convert PDF pages to images and OCR them."""
    try:
        images = convert_from_bytes(file_data, first_page=1, last_page=10, dpi=200)
        pages = []
        for img in images:
            text = pytesseract.image_to_string(img, lang="eng")
            if text.strip():
                pages.append(text.strip())
        return "\n\n".join(pages)
    except Exception:
        return ""
